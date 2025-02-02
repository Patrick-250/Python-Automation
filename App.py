import openpyxl as xl
from openpyxl.chart import BarChart,Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    # cell=sheet1["a1"]
    # print(cell)
    # cell=sheet.cell(1,1)
    # print(cell.value)
    number_of_rows = sheet.max_row
    # print(number_of_rows)
    for row in range(2, number_of_rows + 1):  # loop start from 2 to ignore the headers
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        # print(corrected_price)
    values = Reference(sheet, min_row=2, max_row=sheet.max_row,
                       min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    wb.save("filename")














